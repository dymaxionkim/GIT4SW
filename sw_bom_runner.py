import os
import sys
import gc
import time
import argparse
import json
import ast
import pandas as pd

# Configure stdout and stderr to use UTF-8 and replace encoding errors to avoid crashes with localized chars
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

try:
    import win32com.client
    import pythoncom
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False
    print("Error: pywin32 or pythoncom is not installed.", file=sys.stderr)
    sys.exit(1)

# Import helpers from sw_export_runner
try:
    from sw_export_runner import load_sw_typelib, get_dynamic_sw_app, get_component_model, get_custom_property_value, get_builtin_material, get_builtin_weight
except ImportError:
    # Inline fallback definitions if import fails
    def load_sw_typelib():
        return None
    def get_dynamic_sw_app(raw_obj):
        import win32com.client.dynamic
        return win32com.client.dynamic.Dispatch(raw_obj)
    def get_component_model(comp):
        try:
            return comp.GetModelDoc2()
        except:
            return None
    def get_custom_property_value(prop_mgr, name):
        try:
            res = prop_mgr.Get5(name, False)
            if isinstance(res, tuple) and len(res) >= 3:
                return res[2]
        except:
            pass
        return ""
    def get_builtin_material(model, config_name):
        return ""
    def get_builtin_weight(model):
        return None

def connect_to_solidworks():
    swApp = None
    was_already_running = False
    print("Connecting to SolidWorks...", flush=True)
    try:
        raw_obj = win32com.client.GetActiveObject("SldWorks.Application")
        swApp = get_dynamic_sw_app(raw_obj)
        print("Connected to active SolidWorks instance.", flush=True)
        was_already_running = True
    except Exception:
        pass
        
    if swApp is None:
        sldworks_exe = r"C:\Program Files\SOLIDWORKS Corp\SOLIDWORKS\sldworks.exe"
        if os.path.exists(sldworks_exe):
            print(f"No active instance found. Launching SolidWorks: {sldworks_exe}", flush=True)
            import subprocess
            subprocess.Popen([sldworks_exe])
            poll_timeout = 30.0
            poll_start = time.time()
            while time.time() - poll_start < poll_timeout:
                try:
                    raw_obj = win32com.client.GetActiveObject("SldWorks.Application")
                    swApp = get_dynamic_sw_app(raw_obj)
                    print("Connected to SolidWorks after launching.", flush=True)
                    break
                except Exception:
                    time.sleep(0.5)
                    
    if swApp is None:
        try:
            raw_obj = win32com.client.GetObject(Class="SldWorks.Application")
            swApp = get_dynamic_sw_app(raw_obj)
            print("Connected to SolidWorks via GetObject.", flush=True)
            was_already_running = True
        except Exception as e:
            print(f"Failed to connect to SolidWorks: {e}", file=sys.stderr, flush=True)
            
    return swApp, was_already_running

def save_dataframe_to_excel_with_autowidth(df, filepath):
    print(f"Saving to Excel: {filepath}", flush=True)
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        worksheet = writer.sheets['Sheet1']
        # Build column letter map from DataFrame columns (1-indexed in Excel)
        from openpyxl.utils import get_column_letter
        col_map = {}
        for i, col_name in enumerate(df.columns):
            letter = get_column_letter(i + 1)
            col_map[col_name] = letter

        # Auto-adjust columns width
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # Hide specified columns
        hidden_cols = ["Depth", "Type", "PartNumber", "File Name", "Configuration", "File Path"]
        for col_name in hidden_cols:
            if col_name in col_map:
                worksheet.column_dimensions[col_map[col_name]].hidden = True

        # Set Weight column number format to 6 decimal places
        if "Weight" in col_map:
            weight_letter = col_map["Weight"]
            for row in worksheet.iter_rows(min_col=worksheet[weight_letter][0].column,
                                            max_col=worksheet[weight_letter][0].column):
                for cell in row:
                    cell.number_format = '0.000000'
        
def lookup_property(props, key_name, aliases=None):
    if not props:
        return ""
    # Try direct match
    if key_name in props:
        return props[key_name]
    # Try case-insensitive match
    key_lower = key_name.lower()
    for k, v in props.items():
        if k.lower() == key_lower:
            return v
    # Try aliases
    if aliases:
        for alias in aliases:
            alias_lower = alias.lower()
            for k, v in props.items():
                if k.lower() == alias_lower:
                    return v
    return ""

def main():
    parser = argparse.ArgumentParser(description="SolidWorks Assembly BOM Extractor")
    parser.add_argument("assembly_path", help="Absolute path to the .sldasm file")
    parser.add_argument("--config", help="Specific configuration name to activate", default=None)
    parser.add_argument("--was-running", choices=["true", "false"], default=None, help="Whether SolidWorks was already running before the task started")
    parser.add_argument("--open-before", help="Comma-separated absolute paths of documents open before the BOM button was clicked", default=None)
    args = parser.parse_args()

    assembly_file_path = os.path.abspath(args.assembly_path)
    if not os.path.exists(assembly_file_path):
        print(f"Error: File does not exist at {assembly_file_path}", file=sys.stderr, flush=True)
        sys.exit(1)

    pythoncom.CoInitialize()

    swApp, detected_running = connect_to_solidworks()
    if not swApp:
        print("Error: Could not connect to or start SolidWorks.", file=sys.stderr, flush=True)
        sys.exit(1)

    if args.was_running is not None:
        was_already_running = (args.was_running == "true")
    else:
        was_already_running = detected_running

    model = None
    all_bom_rows = []
    already_open_paths = set()
    try:
        # --- Key fix: already_open_paths is composed only from the --open-before argument ---
        # If refreshed via GetDocuments(), files opened by the config-lookup subprocess
        # (assembly + dependency files) would be added to already_open_paths and could not be closed.
        # --open-before only includes the files the user had open at the moment the BOM button was
        # clicked in the UI, so it is the only reliable source. Without --open-before, use an empty set (=close all).
        if args.open_before:
            for path_str in args.open_before.split(','):
                path_str_clean = path_str.strip()
                if path_str_clean:
                    already_open_paths.add(os.path.normpath(path_str_clean).lower())
            print(f"already_open_paths from --open-before: {len(already_open_paths)} file(s)", flush=True)
            for ap in sorted(already_open_paths):
                print(f"  PRE-EXISTING: {ap}", flush=True)
        else:
            print("No --open-before provided. All open documents will be closed.", flush=True)

        # Load SolidWorks early-bound typelibs
        load_sw_typelib()

        # Open document silently and read-only.
        # swOpenDocOptions_Silent = 1, swOpenDocOptions_ReadOnly = 2,
        # swOpenDocOptions_IgnoreActivationAndSuppression = 64 (suppress automatic loading of parent assembly).
        doc_type = 2 # swDocASSEMBLY
        options = 1 | 2 | 64
        error = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        warning = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)

        print(f"Opening document: {assembly_file_path}", flush=True)
        model = swApp.OpenDoc6(assembly_file_path, doc_type, options, "", error, warning)
        if not model:
            print("Error: Failed to open assembly document in SolidWorks.", file=sys.stderr, flush=True)
            sys.exit(1)

        # Resolve lightweight components
        try:
            print("Resolving lightweight components...", flush=True)
            model.ResolveAllLightWeightComponents(True)
        except Exception as e:
            print(f"Warning resolving lightweight components: {e}", flush=True)

        # If a specific configuration is requested, activate it before extracting the BOM
        if args.config:
            print(f"Activating configuration: {args.config}", flush=True)
            try:
                show_cfg_val = model.ShowConfiguration2
                if callable(show_cfg_val):
                    show_cfg_val(args.config)
                else:
                    model.ShowConfiguration2 = args.config
            except Exception as e_cfg:
                print(f"Warning: Failed to activate configuration '{args.config}': {e_cfg}", flush=True)

        # Get active configuration in a highly robust manner
        config = None
        try:
            cfg_mgr = getattr(model, 'ConfigurationManager', None)
            if cfg_mgr:
                config = getattr(cfg_mgr, 'ActiveConfiguration', None)
        except:
            pass

        if config is None:
            try:
                act_cfg_val = getattr(model, 'GetActiveConfiguration', None)
                if act_cfg_val:
                    config = act_cfg_val() if callable(act_cfg_val) else act_cfg_val
            except:
                pass

        root_comp = None
        if config:
            # Try GetRootComponent3
            try:
                root_val = config.GetRootComponent3
                root_comp = root_val(True) if callable(root_val) else root_val
            except Exception as e:
                print(f"GetRootComponent3 failed: {e}. Trying GetRootComponent2...", flush=True)

            if root_comp is None:
                try:
                    root_val = config.GetRootComponent2
                    root_comp = root_val() if callable(root_val) else root_val
                except Exception as e:
                    print(f"GetRootComponent2 failed: {e}. Trying GetRootComponent...", flush=True)

            if root_comp is None:
                try:
                    root_val = config.GetRootComponent
                    root_comp = root_val() if callable(root_val) else root_val
                except:
                    pass

        if not root_comp:
            print("Error: Could not get the root component of the assembly.", file=sys.stderr, flush=True)
            sys.exit(1)

        all_bom_rows = []
        all_custom_prop_names = set()

        def traverse(comp, depth, parent_absolute_qty):
            try:
                children_val = comp.GetChildren
                children = children_val() if callable(children_val) else children_val
            except Exception as e:
                print(f"Warning: Failed to get children for component: {e}", flush=True)
                return

            if not children:
                return

            # Group children by path & configuration to calculate local quantity at this level
            groups = {}
            for child in children:
                try:
                    # Skip suppressed
                    suppressed_val = child.IsSuppressed
                    is_suppressed = suppressed_val() if callable(suppressed_val) else suppressed_val
                    if is_suppressed:
                        continue
                    
                    # Exclude from BOM check
                    try:
                        if child.ExcludeFromBOM:
                            continue
                    except:
                        pass

                    path_val = child.GetPathName
                    path = path_val() if callable(path_val) else path_val
                    
                    ref_config_val = child.ReferencedConfiguration
                    config_name = ref_config_val() if callable(ref_config_val) else ref_config_val
                    if not path or not config_name:
                        continue

                    key = (path.lower(), config_name.lower())
                    if key not in groups:
                        groups[key] = []
                    groups[key].append(child)
                except Exception as e:
                    print(f"Warning: Failed to process child component: {e}", flush=True)

            # Process groups and recurse
            for key, child_list in groups.items():
                local_qty = len(child_list)
                absolute_qty = parent_absolute_qty * local_qty
                rep_child = child_list[0]

                path_val = rep_child.GetPathName
                path = path_val() if callable(path_val) else path_val
                
                ref_config_val = rep_child.ReferencedConfiguration
                config_name = ref_config_val() if callable(ref_config_val) else ref_config_val

                # Extract properties
                props = {}
                child_model = None
                try:
                    child_model = get_component_model(rep_child)
                except Exception as e:
                    print(f"Warning: Failed to get component model for {path}: {e}", flush=True)

                if child_model:
                    try:
                        # 1. Config-specific custom properties
                        cfg_prop_mgr = child_model.Extension.CustomPropertyManager(config_name)
                        if cfg_prop_mgr:
                            names_val = cfg_prop_mgr.GetNames
                            cfg_names = names_val() if callable(names_val) else names_val
                            if cfg_names:
                                for name in cfg_names:
                                    val = get_custom_property_value(cfg_prop_mgr, name)
                                    props[name] = val
                                    all_custom_prop_names.add(name)
                    except:
                        pass

                    try:
                        # 2. General custom properties fallback
                        gen_prop_mgr = child_model.Extension.CustomPropertyManager("")
                        if gen_prop_mgr:
                            names_val = gen_prop_mgr.GetNames
                            gen_names = names_val() if callable(names_val) else names_val
                            if gen_names:
                                for name in gen_names:
                                    if name not in props or not props[name]:
                                        val = get_custom_property_value(gen_prop_mgr, name)
                                        props[name] = val
                                        all_custom_prop_names.add(name)
                    except:
                        pass

                    # 3. Fallback: if Material is still missing/empty, try built-in material
                    has_material = any(
                        k.lower() == "material" and v and str(v).strip()
                        for k, v in props.items()
                    )
                    if not has_material:
                        mat_source_path = path
                        builtin_mat = get_builtin_material(child_model, config_name)
                        if builtin_mat:
                            props["Material"] = builtin_mat
                            all_custom_prop_names.add("Material")

                    # 4. Fallback: if Weight is missing/empty and file is sldprt/sldasm, try built-in mass
                    if path and (path.lower().endswith('.sldprt') or path.lower().endswith('.sldasm')):
                        has_weight = any(
                            k.lower() in ("weight", "weight(g)", "mass", "weight(kg)") and v and str(v).strip()
                            for k, v in props.items()
                        )
                        if not has_weight:
                            builtin_wt = get_builtin_weight(child_model)
                            if builtin_wt is not None and builtin_wt >= 0:
                                props["Weight"] = f"{builtin_wt:.6f}"
                                all_custom_prop_names.add("Weight")

                    child_model = None

                # Append to raw BOM rows
                all_bom_rows.append({
                    'Depth': depth,
                    'Quantity': local_qty,
                    'AbsoluteQuantity': absolute_qty,
                    'File Path': path,
                    'Configuration': config_name,
                    'Properties': props
                })

                # Recursively traverse the sub-assembly if it is an assembly file (.sldasm)
                if path and path.lower().endswith('.sldasm'):
                    traverse(rep_child, depth + 1, absolute_qty)

        # Start traversal from Depth 1 (children of root)
        print("Traversing assembly components...", flush=True)
        traverse(root_comp, 1, 1)
        print(f"Traversal complete. Found {len(all_bom_rows)} nodes.", flush=True)

        # Load column_order from config.json; fall back to hardcoded default
        default_column_order = [
            "Depth",
            "Type",
            "PartNumber",
            "Partname",
            "Qty",
            "Material",
            "Treatment",
            "Weight",
            "Description",
            "File Name",
            "Configuration",
            "File Path"
        ]
        try:
            config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
            if os.path.exists(config_path):
                with open(config_path, "r", encoding="utf-8") as f:
                    config_data = json.load(f)
                column_order = config_data.get("column_order", default_column_order)
                if isinstance(column_order, str):
                    column_order = ast.literal_eval(column_order)
            else:
                column_order = default_column_order
        except Exception:
            column_order = default_column_order

        def get_mapped_row(depth, qty, file_path, configuration, props, order):
            file_name = os.path.basename(file_path) if file_path else ""
            
            # Fallback for Partname: use base file name if null/empty
            part_name = lookup_property(props, "Partname", ["Part_Name", "Part Name", "Title"])
            if not part_name or str(part_name).strip() == "":
                part_name = os.path.splitext(file_name)[0]
                
            # Prepend {Depth - 1} spaces to Partname for visual hierarchy representation
            spaces_count = max(0, depth - 1)
            part_name = (" " * spaces_count) + str(part_name)
                
            # Type classification (ASM for sldasm, PRT for sldprt)
            ext = os.path.splitext(file_path)[1].lower() if file_path else ""
            if ext == '.sldasm':
                file_type = "ASM"
            elif ext == '.sldprt':
                file_type = "PRT"
            else:
                file_type = ""
                
            part_number = lookup_property(props, "PartNumber", ["PartNo", "Part_No", "Part_Number", "Part Number", "ItemNo", "Item Number"])
            material = lookup_property(props, "Material")
            treatment = lookup_property(props, "Treatment", ["SurfaceTreatment", "Surface Treatment", "Finish"])
            weight = lookup_property(props, "Weight", ["Weight(g)", "Mass", "Weight(kg)"])
            # Normalize weight to 6 decimal places, no scientific notation
            if weight and str(weight).strip():
                try:
                    w = float(str(weight).strip().replace(",", ""))
                    weight = f"{w:.6f}"
                except ValueError:
                    pass
            description = lookup_property(props, "Description", ["Desc"])
            # Fallback: if Description is empty and the part is from Toolbox (SOLIDWORKS Data),
            # use the Configuration value as Description
            if (not description or str(description).strip() == "") and file_path and "SOLIDWORKS Data" in file_path:
                description = configuration

            known = {
                "Depth": depth,
                "Type": file_type,
                "PartNumber": part_number,
                "Partname": part_name,
                "Qty": qty,
                "Material": material,
                "Treatment": treatment,
                "Weight": weight,
                "Description": description,
                "File Name": file_name,
                "Configuration": configuration,
                "File Path": file_path,
            }

            row = {}
            for col in order:
                if col in known:
                    row[col] = known[col]
                else:
                    row[col] = lookup_property(props, col)
            return row

        # 1. Compile Hierarchical BOM Tree DataFrame
        bom_data = []
        for row in all_bom_rows:
            mapped = get_mapped_row(
                depth=row['Depth'],
                qty=row['Quantity'],
                file_path=row['File Path'],
                configuration=row['Configuration'],
                props=row['Properties'],
                order=column_order
            )
            bom_data.append(mapped)

        df_bom = pd.DataFrame(bom_data, columns=column_order)
        if df_bom.empty:
            df_bom = pd.DataFrame(columns=column_order)

        # 2. Compile Aggregated Partlist DataFrame (sum of absolute quantities)
        pl_groups = {}
        for row in all_bom_rows:
            key = (row['File Path'].lower(), row['Configuration'].lower())
            if key not in pl_groups:
                pl_groups[key] = {
                    'Depth': 1,
                    'Quantity': 0,
                    'File Path': row['File Path'],
                    'Configuration': row['Configuration'],
                    'Properties': row['Properties'].copy()
                }
            pl_groups[key]['Quantity'] += row['AbsoluteQuantity']

        pl_data = []
        for key, item in pl_groups.items():
            mapped = get_mapped_row(
                depth=item['Depth'],
                qty=item['Quantity'],
                file_path=item['File Path'],
                configuration=item['Configuration'],
                props=item['Properties'],
                order=column_order
            )
            pl_data.append(mapped)

        df_pl = pd.DataFrame(pl_data, columns=column_order)
        if df_pl.empty:
            df_pl = pd.DataFrame(columns=column_order)

        # BOM: clear Weight for assembly (.sldasm) rows
        if not df_bom.empty and "Type" in df_bom.columns and "Weight" in df_bom.columns:
            df_bom.loc[df_bom["Type"] == "ASM", "Weight"] = ""

        # PL: exclude rows where Type is ASM
        if not df_pl.empty and "Type" in df_pl.columns:
            df_pl = df_pl[df_pl["Type"] != "ASM"].reset_index(drop=True)

        # 3. Create target directory: 2D/BOM relative to assembly path
        target_dir = os.path.dirname(assembly_file_path)
        bom_dir = os.path.join(target_dir, "2D", "BOM")
        os.makedirs(bom_dir, exist_ok=True)

        base_name = os.path.splitext(os.path.basename(assembly_file_path))[0]
        if args.config:
            bom_file_path = os.path.join(bom_dir, f"{base_name}__{args.config}__BOM.xlsx")
            pl_file_path = os.path.join(bom_dir, f"{base_name}__{args.config}__PL.xlsx")
        else:
            bom_file_path = os.path.join(bom_dir, f"{base_name}__BOM.xlsx")
            pl_file_path = os.path.join(bom_dir, f"{base_name}__PL.xlsx")

        # Save files
        save_dataframe_to_excel_with_autowidth(df_bom, bom_file_path)
        save_dataframe_to_excel_with_autowidth(df_pl, pl_file_path)
        print("BOM Tree and Partlist successfully saved to Excel.", flush=True)

    except Exception as e:
        print(f"Error during BOM generation: {e}", file=sys.stderr, flush=True)
        import traceback
        traceback.print_exc()
        sys.exit(1)

    finally:
        # Release COM references first to allow SolidWorks to release document locks
        model = None
        root_comp = None
        config = None
        gc.collect()
        try:
            pythoncom.CoCollectFreeUnusedLibraries()
        except:
            pass

        # Close documents cleanly
        try:
            print(f"Closing documents opened during BOM extraction in SolidWorks...", flush=True)

            # Temporarily set UserControl to False so background closing is not blocked by the UI
            orig_user_control = True
            try:
                orig_user_control = swApp.UserControl
                swApp.UserControl = False
            except Exception:
                pass

            # Path normalization helper: unify slash direction and case to prevent matching failures
            def _norm(p):
                if not p:
                    return ""
                return os.path.normpath(p).replace("\\", "/").lower()

            # Rebuild already_open_paths with the same normalization scheme
            normalized_already_open = set()
            for ap in already_open_paths:
                if ap:
                    normalized_already_open.add(_norm(ap))
                    # Also add the title (file name) form
                    base = os.path.basename(ap)
                    if base:
                        normalized_already_open.add(_norm(base))
                        no_ext = os.path.splitext(base)[0]
                        if no_ext:
                            normalized_already_open.add(no_ext.lower())

            def _is_already_open(path, title):
                """Determine whether path/title was open before the BOM run (normalized matching)"""
                np = _norm(path)
                nt = _norm(title)
                # Full path match
                if np and np in normalized_already_open:
                    return True
                # Title match (both with and without extension)
                if nt:
                    if nt in normalized_already_open:
                        return True
                    no_ext = os.path.splitext(nt)[0]
                    if no_ext and no_ext in normalized_already_open:
                        return True
                return False

            # Close referenced/newly opened files in multiple iterations to resolve dependency locks
            # Assemblies/Drawings must be closed before the parts they reference can be successfully closed.
            last_doc_count = -1
            stuck_count = 0
            for iteration in range(20):
                docs_left = None
                try:
                    val = getattr(swApp, 'GetDocuments')
                    docs_left = val() if callable(val) else val
                except Exception as doc_err:
                    print(f"Warning: Failed to get open documents: {doc_err}", flush=True)
                    break

                if not docs_left:
                    break

                current_count = len(docs_left)
                print(f"BOM cleanup iteration {iteration + 1}: {current_count} document(s) open.", flush=True)
                if current_count == last_doc_count:
                    stuck_count += 1
                    if stuck_count > 5:
                        print(f"BOM cleanup: detected stuck document count ({current_count} docs), breaking to prevent infinite loop.", flush=True)
                        break
                else:
                    stuck_count = 0
                last_doc_count = current_count

                parent_files = []  # list of dict
                child_files = []   # list of dict
                for d in docs_left:
                    try:
                        # Get path name
                        path_val = d.GetPathName
                        path = path_val() if callable(path_val) else path_val
                        if not path:
                            # fallback to title if empty (e.g. unsaved/virtual document)
                            title_val = d.GetTitle
                            title = title_val() if callable(title_val) else title_val
                            path = title

                        title_val = d.GetTitle
                        title = title_val() if callable(title_val) else title_val
                                
                        # If not a document open before the BOM run (= newly opened), it is a close target.
                        is_pre_existing = _is_already_open(path, title)
                        if not is_pre_existing:
                            try:
                                dtype = d.GetType()
                            except:
                                dtype = getattr(d, 'GetType')
                                if callable(dtype):
                                    dtype = dtype()
                            path_lower = (path or "").lower()
                            title_lower = (title or "").lower()
                            is_parent = (dtype in (2, 3) or 
                                         path_lower.endswith(".sldasm") or 
                                         path_lower.endswith(".slddrw") or
                                         title_lower.endswith(".sldasm") or
                                         title_lower.endswith(".slddrw"))
                            
                            # Generate unique closing identifiers
                            ids = []
                            if path:
                                ids.append(path)
                                ids.append(os.path.normpath(path))
                                base = os.path.basename(path)
                                if base:
                                    ids.append(base)
                                    base_no_ext = os.path.splitext(base)[0]
                                    if base_no_ext:
                                        ids.append(base_no_ext)
                            if title:
                                ids.append(title)
                                title_no_ext = os.path.splitext(title)[0]
                                if title_no_ext:
                                    ids.append(title_no_ext)
                            
                            uniq_ids = []
                            for iid in ids:
                                if iid and isinstance(iid, str):
                                    iid_s = iid.strip()
                                    if iid_s and iid_s not in uniq_ids:
                                        uniq_ids.append(iid_s)
                                        
                            doc_entry = {
                                'title': title,
                                'path': path,
                                'uniq_ids': uniq_ids
                            }
                            if is_parent:
                                parent_files.append(doc_entry)
                            else:
                                child_files.append(doc_entry)
                        else:
                            # For documents open before the BOM run, only log
                            print(f"  Skipping (pre-existing): {title} (path: {path})", flush=True)
                    except Exception as e_inspect:
                        print(f"Warning: Failed to inspect document for cleanup: {e_inspect}", flush=True)

                # Clear COM references to docs_left before closing to avoid lock
                docs_left = None
                d = None
                gc.collect()
                try:
                    pythoncom.CoCollectFreeUnusedLibraries()
                except:
                    pass

                if not parent_files and not child_files:
                    break

                print(f"  Closing {len(parent_files)} parent(s) + {len(child_files)} child(ren)...", flush=True)

                closed_any = False
                # Close parents first to release references on children
                for doc_entry in parent_files:
                    title = doc_entry['title']
                    path = doc_entry['path']
                    uniq_ids = doc_entry['uniq_ids']
                    print(f"Closing newly opened/referenced parent document: {title} (path: {path})", flush=True)
                    for identifier in uniq_ids:
                        try:
                            swApp.CloseDoc(identifier)
                            closed_any = True
                            print(f"  Closed via CloseDoc('{identifier}')", flush=True)
                            break
                        except Exception as e_close_parent:
                            pass
                        try:
                            swApp.QuitDoc(identifier)
                            closed_any = True
                            print(f"  Closed via QuitDoc('{identifier}')", flush=True)
                            break
                        except Exception as e_quit_parent:
                            pass
                    else:
                        print(f"  ⚠️ Failed to close parent: {title} (path: {path})", flush=True)

                if closed_any:
                    time.sleep(0.15)

                # Close children second
                for doc_entry in child_files:
                    title = doc_entry['title']
                    path = doc_entry['path']
                    uniq_ids = doc_entry['uniq_ids']
                    print(f"Closing newly opened/referenced child document: {title} (path: {path})", flush=True)
                    for identifier in uniq_ids:
                        try:
                            swApp.CloseDoc(identifier)
                            closed_any = True
                            print(f"  Closed via CloseDoc('{identifier}')", flush=True)
                            break
                        except Exception as e_close_child:
                            pass
                        try:
                            swApp.QuitDoc(identifier)
                            closed_any = True
                            print(f"  Closed via QuitDoc('{identifier}')", flush=True)
                            break
                        except Exception as e_quit_child:
                            pass
                    else:
                        print(f"  ⚠️ Failed to close child: {title} (path: {path})", flush=True)

                time.sleep(0.15)
                # Release COM references then run garbage collection to release locks
                gc.collect()
                try:
                    pythoncom.CoCollectFreeUnusedLibraries()
                except:
                    pass

            # Restore UserControl
            try:
                swApp.UserControl = orig_user_control
            except Exception:
                pass

            # ----- Final verification pass -----
            # Check whether any documents are still open after the close loop completes.
            # If any newly opened documents (not in already_open_paths) remain, the close
            # command failed; emit a warning and retry closing.
            for verify_iter in range(3):
                docs_remaining = None
                try:
                    val = getattr(swApp, 'GetDocuments')
                    docs_remaining = val() if callable(val) else val
                except:
                    pass

                if not docs_remaining:
                    print("BOM cleanup verified: all newly opened documents are closed.", flush=True)
                    break

                # Collection target: documents newly opened during the BOM run that are still open
                still_open = []
                for d in docs_remaining:
                    try:
                        path_val = d.GetPathName
                        path = path_val() if callable(path_val) else path_val
                        if not path:
                            tv = d.GetTitle
                            path = tv() if callable(tv) else tv
                        tv2 = d.GetTitle
                        title = tv2() if callable(tv2) else tv2

                        is_pre_existing = _is_already_open(path, title)
                        if not is_pre_existing:
                            still_open.append({'title': title, 'path': path})
                    except Exception as ie:
                        print(f"Warning: Failed to inspect remaining document: {ie}", flush=True)

                if not still_open:
                    print(f"BOM cleanup verified: {len(docs_remaining)} document(s) remain open but all were open before BOM run.", flush=True)
                    break

                # Emit warnings for documents that were attempted to close but failed
                for entry in still_open:
                    print(f"⚠️ Warning: document failed to close and is still open: '{entry['title']}' (path: {entry['path']})", flush=True)
                print(f"Verification retry {verify_iter + 1}/3: {len(still_open)} newly opened document(s) still open. Retrying close...", flush=True)

                # Close parents (asm/drw) first
                parents = [e for e in still_open if (e['path'] or "").lower().endswith((".sldasm", ".slddrw")) or (e['title'] or "").lower().endswith((".sldasm", ".slddrw"))]
                children = [e for e in still_open if e not in parents]

                def _try_close(entry):
                    ids = []
                    p = entry['path']
                    t = entry['title']
                    if p:
                        ids.append(p)
                        b = os.path.basename(p)
                        if b:
                            ids.append(b)
                            ids.append(os.path.splitext(b)[0])
                    if t:
                        ids.append(t)
                        ids.append(os.path.splitext(t)[0])
                    seen = []
                    for i in ids:
                        if i and isinstance(i, str) and i.strip() and i.strip() not in seen:
                            seen.append(i.strip())
                    closed = False
                    for identifier in seen:
                        try:
                            swApp.CloseDoc(identifier)
                            print(f"  Retry CloseDoc('{identifier}') OK", flush=True)
                            closed = True
                            break
                        except Exception as e1:
                            pass
                        try:
                            swApp.QuitDoc(identifier)
                            print(f"  Retry QuitDoc('{identifier}') OK", flush=True)
                            closed = True
                            break
                        except Exception as e2:
                            pass
                    return closed

                closed_any = False
                for e in parents:
                    if _try_close(e):
                        closed_any = True
                if closed_any:
                    time.sleep(0.15)
                closed_any2 = False
                for e in children:
                    if _try_close(e):
                        closed_any2 = True
                if closed_any or closed_any2:
                    time.sleep(0.15)

            # Final re-check: if any newly opened documents still remain, emit a clear error message
            docs_final = None
            try:
                val = getattr(swApp, 'GetDocuments')
                docs_final = val() if callable(val) else val
            except:
                pass

            if docs_final:
                leftover = []
                for d in docs_final:
                    try:
                        pv = d.GetPathName
                        p = pv() if callable(pv) else pv
                        if not p:
                            tv = d.GetTitle
                            p = tv() if callable(tv) else tv
                        tv2 = d.GetTitle
                        t = tv2() if callable(tv2) else tv2
                        if not _is_already_open(p, t):
                            leftover.append((t, p))
                    except:
                        pass
                if leftover:
                    print("Error: BOM cleanup FAILED - the following newly opened documents could not be closed:", flush=True)
                    for t, p in leftover:
                        print(f"Error:   - '{t}' (path: {p})", flush=True)
                else:
                    print(f"BOM cleanup OK: {len(docs_final)} pre-existing document(s) remain open (expected).", flush=True)
            else:
                print("BOM cleanup OK: no documents remain open.", flush=True)
        except Exception as e_close_all:
            print(f"Warning: Error during referenced docs cleanup: {e_close_all}", file=sys.stderr, flush=True)

        # Close SolidWorks if we launched it and it wasn't already running before the script started
        if swApp:
            if not was_already_running:
                import threading
                def exit_sw_async(app):
                    try:
                        print("Exiting SolidWorks (launched by BOM runner)...", flush=True)
                        app.ExitApp()
                    except Exception as exit_e:
                        print(f"Warning: Error during swApp.ExitApp(): {exit_e}", flush=True)
                        
                try:
                    t_exit = threading.Thread(target=exit_sw_async, args=(swApp,), daemon=True)
                    t_exit.start()
                    t_exit.join(timeout=5.0)
                except Exception as t_err:
                    print(f"Failed to start async ExitApp thread: {t_err}", flush=True)
            else:
                print("SolidWorks was already running before BOM extraction. Keeping SolidWorks open.", flush=True)

        gc.collect()
        
        # Force terminate SolidWorks processes if we started it and it is still alive
        if not was_already_running:
            try:
                import subprocess
                subprocess.run("taskkill /F /IM SLDWORKS.exe", shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                subprocess.run("taskkill /F /IM sldworks_fs.exe", shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                print("Forcefully terminated remaining SolidWorks processes spawned by BOM runner.", flush=True)
            except Exception as kill_e:
                print(f"Could not clean up SolidWorks processes: {kill_e}", flush=True)

        pythoncom.CoUninitialize()

if __name__ == "__main__":
    main()
